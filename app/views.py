from flask import render_template, flash, redirect, session, url_for, request, \
    g, jsonify, send_from_directory
from flask_login import login_user, logout_user, current_user, login_required
from flask_sqlalchemy import get_debug_queries
from werkzeug import secure_filename
from flask_babel import gettext
from datetime import datetime
from dateutil.parser import parse
from guess_language import guessLanguage
from app import app, db, lm, babel
from .forms import EditForm, PostForm, SearchForm, RegistrationForm, LoginForm, \
    ManpowerForm, MobilityForm
from .models import User, Post, Mobility, Rules
from .emails import follower_notification
from .translate import microsoft_translate
from .helpers import *
from config import POSTS_PER_PAGE, MAX_SEARCH_RESULTS, LANGUAGES, \
    DATABASE_QUERY_TIMEOUT, basedir, ALLOWED_EXTENSIONS


@app.route('/mobility_cbt_manager',methods = ['GET', 'POST'])
@login_required
def mobility_cbt_manager():
    header=Rules.query.filter(Rules.name.like('%cbt%')).all()
    if request.method=='POST':
        print(request.form)
        for head in header:
            if request.form.get(head.name) != '' and request.form.get(head.name)!='None':
                if request.form.get(head.name)=='Delete':
                    head.args=None
                else:
                    head.args=request.form.get(head.name)                    
        db.session.commit()
        #Rebuild after update
        header=Rules.query.filter(Rules.name.like('%cbt%')).all()
    return render_template('mobility_cbt_manager.html', header=header)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/mobility_import_html')
@login_required
def mobility_import_html():
    return render_template('mobility_import_html.html')


@app.route('/uploads/<filename>')
def uploaded_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

@app.route('/uploader', methods = ['GET', 'POST'])
def uploader():
    if request.method == 'POST':
        # check if the post request has the file part
        if 'file' not in request.files:
            flash('No file part')
            return redirect(request.url)
        file = request.files['file']
        # if user does not select file, browser also
        # submit a empty part without filename
        if file.filename == '':
            flash('No selected file')
            return redirect(request.url)
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            cbt(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            #flash("File " + str(file.filename) + " Uploaded")
    user=Mobility.query.filter_by(username=g.user.username).first()
    return render_template('mobility_import_html.html')


@app.route('/mobility', methods=['GET', 'POST'])
@login_required
def mobility():
    user=Mobility.query.filter_by(username=g.user.username).first()
    return render_template('mobility.html', title='Mobility', user=user, now=str(datetime.now()))


@app.route('/mobility_query', methods=['GET', 'POST'])
@login_required
def mobility_query():
    mobility=Mobility.query.order_by(Mobility.username).all()
    header=[]
    for head in Mobility.__table__.columns:
        header.append(head)
    return render_template('mobility_query.html',title='Mobility Query',
        mobility=mobility, header=header, now=str(datetime.now()))


@app.route('/mobility_backup', methods=['GET', 'POST'])
@login_required
def mobility_backup():
    from openpyxl import Workbook
    if request.method=='POST':
        if request.form['submit']=='SaveBackup':
            wb = Workbook()
            ws = wb.active
            mobility=Mobility.query.order_by(Mobility.username).all()
            header=[]
            columnC=0
            for head in Mobility.__table__.columns:
                header.append(head)
                columnC=columnC+1
                ws.cell(column=columnC, row=1, value=head.name)
            #Header starts at row=1, this makes loop start at row=2. openpyxl does not take 0
            row=1
            for user in mobility:
                row=row+1
                columnC=0
                for head in header:
                    columnC=columnC+1
                    ws.cell(column=columnC, row=row, value=getattr(user,head.name))
            wb.save("backup/mobility@"+str(datetime.now())+".xlsx")
            flash("Backup "+str(datetime.now())+".xlsx saved")
        elif request.form['submit']=='RestoreFromBackup':
            import os
            files=os.listdir("backup")
            dates=[]
            for date in files:
                dates.append(date.split('@')[1])
    return render_template('/mobility_backup.html',title='Mobility Backup')


@app.route('/mobility_edit_user', methods=['GET', 'POST'])
@login_required
def mobility_edit_user():
    print("EXECUTING!!!")
    #print(request.form)
    #build user list
    data=User.query.with_entities(User.username).all()
    users=[]
    for i in data:
        users.append(i[0])
    #make header
    header=[]
    for head in Mobility.__table__.columns:
        if head.name != 'username':
            header.append(head)
    if request.method == "POST":
        print("POSTING")
        try:
            request.form['submitb']
            u=Mobility.query.filter_by(username=request.form.get('user')).first()
            for head in header:
                if request.form.get(head.name) != '' and request.form.get(head.name)!='None':
                    if request.form.get(head.name)=='Delete':
                        setattr(u,head.name,None)
                    else:
                        if str(head.type) == 'DATETIME':
                            try:
                                date=parse(request.form.get(head.name))
                                setattr(u, head.name, date)
                            except:
                                flash(request.form.get(head.name) + " is not a valid date for " + head.name)                            
                        elif str(head.type) == 'BOOLEAN':
                            setattr(u, head.name,request.form.get(head.name) in ['True', 'T','true','t'])
                        elif str(head.type) == 'INTEGER':
                            try:
                                setattr(u, head.name, int(request.form.get(head.name)))
                            except:
                                flash(request.form.get(head.name) + " is not a valid integer for " + head.name)
                        elif str(head.type) == 'NUMERIC':
                            try:
                                setattr(u, head.name, float(request.form.get(head.name)))
                            except:
                                flash(request.form.get(head.name) + " is not a valid number for " + head.name)
                        elif str(head.type) == 'STRING':
                            setattr(u, head.name, request.form.get(head.name))
                        elif str(head.type).split('(')[0] == 'VARCHAR':
                            setattr(u, head.name, request.form.get(head.name))
                        else:
                            flash("Unknow data type" + str(head.type) + "in variable " + head.name)
            db.session.commit()
            #puts selected user in front of users
            users.insert(0, users.pop(users.index(request.form.get('user'))))
            data=Mobility.query.filter_by(username=request.form.get('user')).first()
            flash("Updated User " + request.form.get('user'))
        except:
            #moves selected user to front of users list
            users.insert(0, users.pop(users.index(request.form.get('user'))))
            #Pulls user data from database
            data=Mobility.query.filter_by(username=request.form.get('user')).first()
            flash("Selected User " + request.form.get('user'))
    else:
        #Pulls user data from database
        data=Mobility.query.filter_by(username=users[0]).first()
        flash("Selected User " + users[0])
    return render_template('mobility_edit_user.html',title='Mobility Edit User', 
        data=data, header=header, users=users)


@app.route('/manpower', methods=['GET', 'POST'])
@login_required
def manpower():
    return render_template('manpower.html',title='Manpower')


@app.route('/manpower_query', methods=['GET', 'POST'])
@login_required
def manpower_query():
    users=User.query.all()
    header=[]
    for head in User.__table__.columns:
        if head.name != 'password_hash' and head.name != 'id':
            header.append(head)
    return render_template('manpower_query.html',title='Manpower Query',
        users=users, header=header)


@app.route('/manpower_add_users', methods=['GET', 'POST'])
@login_required
def manpower_add_users():
    form = RegistrationForm()
    if request.method == 'POST' and form.validate():
        if User.query.filter_by(username=form.username.data).first() == None:
            user = User(username=form.username.data.lower())
            user.email = form.email.data
            user.hash_password(form.password.data)
            db.session.add(user)
            # add username to Mobility data base
            userM=Mobility(username=form.username.data.lower())
            db.session.add(userM)
            # make the user follow him/herself
            db.session.add(user.follow(user))
            db.session.commit()
            flash('User added')
            return redirect(url_for('manpower_add_users'))
        else:
            flash('Username already exists')
            return render_template('register.html', form=form)
    return render_template('manpower_add_users.html', form=form)


@app.route('/manpower_edit_users', methods=['GET', 'POST'])
@login_required
def manpower_edit_users():
    #build user list
    data=User.query.with_entities(User.username).all()
    users=[]
    for i in data:
        users.append(i[0])
    #make header
    header=[]
    for head in User.__table__.columns:
        if head.name != 'password_hash' and head.name != 'id':
            header.append(head)
    if request.method == "POST":
        try:
            request.form['submitb']
            u=User.query.filter_by(username=request.form.get('user')).first()
            for head in header:
                if request.form.get(head.name) != '' and request.form.get(head.name) != 'None':
                    if request.form.get(head.name)=='Delete':
                        setattr(u,head.name,None)
                    else:
                        if str(head.type) == 'DATETIME':
                            try:
                                date=parse(request.form.get(head.name))
                                setattr(u, head.name, date)
                            except:
                                flash(request.form.get(head.name) + " is not a valid date for " + head.name)                            
                        elif str(head.type) == 'BOOLEAN':
                            setattr(u, head.name,request.form.get(head.name) in ['True', 'T','true','t'])
                        elif str(head.type) == 'INTEGER':
                            try:
                                setattr(u, head.name, int(request.form.get(head.name)))
                            except:
                                flash(request.form.get(head.name) + " is not a valid integer for " + head.name)
                        elif str(head.type) == 'NUMERIC':
                            try:
                                setattr(u, head.name, float(request.form.get(head.name)))
                            except:
                                flash(request.form.get(head.name) + " is not a valid number for " + head.name)
                        elif str(head.type) == 'STRING':
                            setattr(u, head.name, request.form.get(head.name))
                        elif str(head.type).split('(')[0] == 'VARCHAR':
                            setattr(u, head.name, request.form.get(head.name))
                        else:
                            flash("Unknow data type" + str(head.type) + "in variable " + head.name)
            try:
                db.session.commit()
                flash("Updated User " + request.form.get('user'))
                
            except:
                flash("Unable to update user + request.form.get('user')")
            users.insert(0, users.pop(users.index(request.form.get('user'))))
            data=User.query.filter_by(username=request.form.get('user')).first()
        except:
            #moves selected user to front of users list
            users.insert(0, users.pop(users.index(request.form.get('user'))))
            #Pulls user data from database
            data=User.query.filter_by(username=request.form.get('user')).first()
            flash("Selected User " + request.form.get('user'))
    else:
        #Pulls user data from database
        data=User.query.filter_by(username=users[0]).first()
        flash("Selected User " + users[0])
    return render_template('manpower_edit_users.html',title='Manpower Edit User', 
        data=data, header=header, users=users)


@app.route('/', methods=['GET', 'POST'])
@app.route('/index', methods=['GET', 'POST'])
@app.route('/index/<int:page>', methods=['GET', 'POST'])
@login_required
def index(page=1):
    form = PostForm()
    if form.validate_on_submit():
        language = guessLanguage(form.post.data)
        if language == 'UNKNOWN' or len(language) > 5:
            language = ''
        post = Post(body=form.post.data, timestamp=datetime.utcnow(),
                    author=g.user, language=language)
        db.session.add(post)
        db.session.commit()
        flash(gettext('Your post is now live!'))
        return redirect(url_for('index'))
    posts = g.user.followed_posts().paginate(page, POSTS_PER_PAGE, False)
    return render_template('index.html',
                           title='Home',
                           form=form,
                           posts=posts)


@app.route('/user/<username>')
@app.route('/user/<username>/<int:page>')
@login_required
def user(username, page=1):
    user = User.query.filter_by(username=username).first()
    if user is None:
        flash(gettext('User %(username)s not found.', username=username))
        return redirect(url_for('index'))
    posts = user.posts.paginate(page, POSTS_PER_PAGE, False)
    return render_template('user.html',
                           user=user,
                           posts=posts,
               title='Your Profile')

@app.route('/edit', methods=['GET', 'POST'])
@login_required
def edit():
    form = EditForm(g.user.username)
    if form.validate_on_submit():
        g.user.username = form.username.data
        g.user.about_me = form.about_me.data
        g.user.hash_password(form.password.data)
        m=Mobility.query.filter_by(username=g.user.username).first()
        m.username=form.username.data
        db.session.add(g.user)
        db.session.commit()
        flash(gettext('Your changes have been saved.'))
        return redirect(url_for('edit'))
    elif request.method != "POST":
        form.username.data = g.user.username
        form.about_me.data = g.user.about_me
    return render_template('edit.html', form=form)


@babel.localeselector
def get_locale():
    return request.accept_languages.best_match(LANGUAGES.keys())


@app.errorhandler(404)
def not_found_error(error):
    return render_template('404.html'), 404


@app.errorhandler(500)
def internal_error(error):
    db.session.rollback()
    return render_template('500.html'), 500


@app.route('/follow/<username>')
@login_required
def follow(username):
    user = User.query.filter_by(username=username).first()
    if user is None:
        flash('User %s not found.' % username)
        return redirect(url_for('index'))
    if user == g.user:
        flash(gettext('You can\'t follow yourself!'))
        return redirect(url_for('user', username=username))
    u = g.user.follow(user)
    if u is None:
        flash(gettext('Cannot follow %(username)s.', username=username))
        return redirect(url_for('user', username=username))
    db.session.add(u)
    db.session.commit()
    flash(gettext('You are now following %(username)s!', username=username))
    follower_notification(user, g.user)
    return redirect(url_for('user', username=username))


@app.route('/unfollow/<username>')
@login_required
def unfollow(username):
    user = User.query.filter_by(username=username).first()
    if user is None:
        flash('User %s not found.' % username)
        return redirect(url_for('index'))
    if user == g.user:
        flash(gettext('You can\'t unfollow yourself!'))
        return redirect(url_for('user', username=username))
    u = g.user.unfollow(user)
    if u is None:
        flash(gettext('Cannot unfollow %(username)s.', username=username))
        return redirect(url_for('user', username=username))
    db.session.add(u)
    db.session.commit()
    flash(gettext('You have stopped following %(username)s.',
                  username=username))
    return redirect(url_for('user', username=username))


@app.route('/delete/<int:id>')
@login_required
def delete(id):
    post = Post.query.get(id)
    if post is None:
        flash('Post not found.')
        return redirect(url_for('index'))
    if post.author.id != g.user.id:
        flash('You cannot delete this post.')
        return redirect(url_for('index'))
    db.session.delete(post)
    db.session.commit()
    flash('Your post has been deleted.')
    return redirect(url_for('index'))


@app.route('/search', methods=['POST'])
@login_required
def search():
    if not g.search_form.validate_on_submit():
        return redirect(url_for('index'))
    return redirect(url_for('search_results', query=g.search_form.search.data))


@app.route('/translate', methods=['POST'])
@login_required
def translate():
    return jsonify({
        'text': microsoft_translate(
            request.form['text'],
            request.form['sourceLang'],
            request.form['destLang'])})


@lm.user_loader
def load_user(id):
    return User.query.get(int(id))


@app.before_request
def before_request():
    g.user = current_user
    if g.user.is_authenticated:
        g.user.last_seen = datetime.utcnow()
        db.session.add(g.user)
        db.session.commit()
        g.search_form = SearchForm()
    g.locale = get_locale()


@app.after_request
def after_request(response):
    for query in get_debug_queries():
        if query.duration >= DATABASE_QUERY_TIMEOUT:
            app.logger.warning(
                "SLOW QUERY: %s\nParameters: %s\nDuration: %fs\nContext: %s\n" %
                (query.statement, query.parameters, query.duration,
                 query.context))
    return response


@app.route('/login', methods=['GET', 'POST'])
def login():
    if g.user is not None and g.user.is_authenticated:
        return redirect(url_for('index'))
    form=LoginForm()
    if form.validate_on_submit():
        session['remember_me'] = form.remember_me.data
        user=User.query.filter_by(username=form.username.data.lower()).first()
        if user and user.verify_password(form.password.data):
            if 'remember_me' in session:
                remember_me = session['remember_me']
                session.pop('remember_me', None)
            login_user(user, remember=remember_me)
            flash("Logged in sucessfully.")
            '''eventually update to add security to redirects
            if not is_safe_url(next):
                return flask.abort(400)
            return redirect(request.args.get('next') or url_for('index'))
            '''
            return redirect(request.args.get('next') or url_for('index'))
        flash("Incorrect username or password")
    return render_template('login.html',
                           title='Sign In',
                           form=form)


@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('index'))

''' old with wtforms implementation
@app.route('/manpower_edit_users', methods=['GET', 'POST'])
@login_required
def manpower_edit_users():
    #build user list
    data=User.query.with_entities(User.username).all()
    users=[]
    for i in data:
        users.append(i[0])

    #for attr, value in form._fields.iteritems():
    #    print(attr, value.error)

    form=ManpowerForm(request.form)
    if request.method == "POST":
        if request.form['submit']=='SelectUser':
            print("SelectUser")
            #moves selected user to front of users list
            users.insert(0, users.pop(users.index(request.form.get('user'))))
            #Pulls user data from database
            data=User.query.filter_by(username=request.form.get('user')).first()
            #PROBLEM CODE does not take in data from object at all
            form=ManpowerForm(obj=data)
            flash("Selected User " + request.form.get('user'))
        elif request.form['submit']=='Update' and form.validate():
            u=User.query.filter_by(username=request.form.get('user')).first()
            form.populate_obj(u)
            db.session.commit()
            #puts selected user in front of users
            users.insert(0, users.pop(users.index(request.form.get('user'))))
            data=User.query.filter_by(username=request.form.get('user')).first()
            form=ManpowerForm(obj=data)
            flash("Updated User " + request.form.get('user'))
        else:
            #move selected user to front of form
            users.insert(0, users.pop(users.index(request.form.get('user'))))
            data=User.query.filter_by(username=request.form.get('user')).first()
            form=ManpowerForm(obj=data)
            flash("Form not validated")
    elif request.method == "GET":
        flash("GET")
    return render_template('manpower_edit_users.html',title='Manpower Edit User', 
        form=form, users=users)
'''

'''
@app.route('/mobility_edit_user', methods=['GET', 'POST'])
@login_required
def mobility_edit_user():
    #build user list
    data=User.query.with_entities(User.username).all()
    users=[]
    for i in data:
        users.append(i[0])
    form=MobilityForm(request.form)
    print(dir(form))
    print(form.meta)
    if request.method == "POST":
        if request.form['submit']=='SelectUser':
            print("SelectUser")
            #moves selected user to front of users list
            users.insert(0, users.pop(users.index(request.form.get('user'))))
            #Pulls user data from database
            data=Mobility.query.filter_by(username=request.form.get('user')).first()
            #PROBLEM CODE does not take in data from object at all
            form=MobilityForm(obj=data)
            flash("Selected User " + request.form.get('user'))
        elif request.form['submit']=='Update' and form.validate():
            u=Mobility.query.filter_by(username=request.form.get('user')).first()
            form.populate_obj(u)
            db.session.commit()
            #puts selected user in front of users
            users.insert(0, users.pop(users.index(request.form.get('user'))))
            data=Mobility.query.filter_by(username=request.form.get('user')).first()
            form=MobilityForm(obj=data)
            flash("Updated User " + request.form.get('user'))
        else:
            #move selected user to front of form
            users.insert(0, users.pop(users.index(request.form.get('user'))))
            data=Mobility.query.filter_by(username=request.form.get('user')).first()
            form=MobilityForm(obj=data)
            flash("Form not validated")
    elif request.method == "GET":
        flash("GET")
    return render_template('mobility_edit_user.html',title='Mobility Edit User', 
        form=form, users=users)
'''